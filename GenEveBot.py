from PyPDF2 import PdfReader
from langchain.embeddings.openai import OpenAIEmbeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.chains import ConversationalRetrievalChain
import pickle
from dotenv import load_dotenv
import os
import streamlit as st
from streamlit_chat import message
import io
import asyncio
import openai
import docx
import openpyxl
import csv
from PIL import Image

# Set the page configuration for the Streamlit app
st.set_page_config(page_title="GenEveBot", page_icon="GenEveBot")

# Set the OpenAI API key
openai.api_key = "Openai_key"

# Load environment variables from .env file
load_dotenv()
api_key = "Openai_key"


async def main():
    # Define the available bot options
    activities = ["Select Bot", "GenBot", "EveBot"]
    choice = st.sidebar.selectbox("Select Bot", activities)


    # If the user selects "Select Bot"
    if choice == "Select Bot":
        # Load and display the image
        img = Image.open("chatbotimg.png")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.write("")
        with col2:
            st.image(img, use_column_width=True)
        with col3:
            st.write("")

        # Display the welcome message
        title_alignment = """<h1 style='text-align: center; color: black;'>Welcome to Gen & Eve Bot!</h1>"""
        st.markdown(title_alignment, unsafe_allow_html=True)

    # If the user selects "GenBot"
    elif choice == 'GenBot':
        # Function to store document embeddings
        async def storeDocEmbeds(file, filename):
            # Handle PDF files
            if filename.endswith(".pdf"):
                reader = PdfReader(file)
                corpus = ''.join([p.extract_text() for p in reader.pages if p.extract_text()])
                splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
                chunks = splitter.split_text(corpus)
                embeddings = OpenAIEmbeddings(openai_api_key=api_key)
                vectors = FAISS.from_texts(chunks, embeddings)
                with open(filename + ".pkl", "wb") as f:
                    pickle.dump(vectors, f)
            
            # Handle DOCX files
            elif filename.endswith(".docx"):
                doc = docx.Document(file)
                fullText = []
                for para in doc.paragraphs:
                    fullText.append(para.text)
                corpus = ' '.join(fullText)
                splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
                chunks = splitter.split_text(corpus)
                embeddings = OpenAIEmbeddings(openai_api_key=api_key)
                vectors = FAISS.from_texts(chunks, embeddings)
                with open(filename + ".pkl", "wb") as f:
                    pickle.dump(vectors, f)
            
            # Handle XLSX files
            elif filename.endswith(".xlsx"):
                workbook = openpyxl.load_workbook(file)
                corpus = ''
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    for row in worksheet.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value:
                                corpus += str(cell_value) + ' '
                splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
                chunks = splitter.split_text(corpus)
                embeddings = OpenAIEmbeddings(openai_api_key=api_key)
                vectors = FAISS.from_texts(chunks, embeddings)
                with open(filename + ".pkl", "wb") as f:
                    pickle.dump(vectors, f)
            
            # Handle CSV files
            elif filename.endswith(".csv"):
                content = file.getvalue().decode('utf-8')
                reader = csv.reader(content.splitlines())
                corpus = ''
                for row in reader:
                    for cell in row:
                        corpus += str(cell) + ' '
                splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
                chunks = splitter.split_text(corpus)
                embeddings = OpenAIEmbeddings(openai_api_key=api_key)
                vectors = FAISS.from_texts(chunks, embeddings)
                with open(filename + ".pkl", "wb") as f:
                    pickle.dump(vectors, f)

        # Function to get document embeddings
        async def getDocEmbeds(file, filename):
            if not os.path.isfile(filename + ".pkl"):
                await storeDocEmbeds(file, filename)
            with open(filename + ".pkl", "rb") as f:
                global vectors
                vectors = pickle.load(f)
            return vectors

        # Function for conversational chat
        async def conversational_chat(query):
            result = qa({"question": query, "chat_history": st.session_state['history']})
            st.session_state['history'].append((query, result["answer"]))
            return result["answer"]

        # Check if chat history exists in session state, if not, initialize it
        if 'history' not in st.session_state:
            st.session_state['history'] = []

        # Define the GenBot interface
        label = Image.open("chatbotimg.png")
        new_image1 = label.resize((2000, 2000))
        col1, col2, col3 = st.columns([1, 6, 1])
        with col1:
            st.image(new_image1, use_column_width=None)
        with col2:
            st.title("GenBot")
        with col3:
            st.write("")

        # Check if the app is ready to process files
        if 'ready' not in st.session_state:
            st.session_state['ready'] = False

        # Display file uploader
        uploaded_file = st.file_uploader("Choose a file", type=["pdf", "docx", "xlsx", "csv"])

        # Process the uploaded file
        if uploaded_file:
            st.session_state['ready'] = False  # Reset ready state
            with st.spinner("Processing..."):
                file_content = uploaded_file.read()
                vectors = await getDocEmbeds(io.BytesIO(file_content), uploaded_file.name)
                # Initialize QA model
                qa = ConversationalRetrievalChain.from_llm(
                    ChatOpenAI(model_name="gpt-3.5-turbo", openai_api_key=api_key),
                    retriever=vectors.as_retriever(),
                    return_source_documents=True
                )

                st.session_state['ready'] = True



        st.divider()

        if st.session_state['ready']:
            if 'generated' not in st.session_state:
                st.session_state['generated'] = ["Welcome! You can now ask any questions regarding " + uploaded_file.name]

            if 'past' not in st.session_state:
                st.session_state['past'] = ["Hi Bot!"]

            # Container for chat history
            response_container = st.container()

            # Container for text box
            container = st.container()

            with container:
                with st.form(key='my_form', clear_on_submit=True):
                    user_input = st.text_input("Ask something:", placeholder="e.g: Summarize the paper in a few sentences", key='input')
                    submit_button = st.form_submit_button(label='Send')

                if submit_button and user_input:
                    output = await conversational_chat(user_input)
                    st.session_state['past'].append(user_input)
                    st.session_state['generated'].append(output)

            if st.session_state['generated']:
                with response_container:
                    for i in range(len(st.session_state['generated'])):
                        message(st.session_state["past"][i], is_user=True, key=str(i) + '_user', avatar_style="thumbs")
                        message(st.session_state["generated"][i], key=str(i) + '🤖')
                        # , avatar_style="thumbs")👽🐧

    # Implementation for EveBot
    elif choice == 'EveBot':
        label = Image.open("chatbotimg.png")
        new_image1 = label.resize((2000, 2000))
        col1, col2, col3 = st.columns([1, 6, 1])
        with col1:
            st.image(new_image1, use_column_width=None)
        with col2:
            st.title("EveBot")
        with col3:
            st.write("")

        # Set OpenAI API key
        os.environ["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"] = "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"
        openai.api_key = os.getenv("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")

        st.session_state['ready'] = True

        # Function to generate response from EveBot
        async def res(input):
            response = openai.Completion.create(
                model="text-davinci-003",
                prompt="Q:" + input + "\n " + "A:",
                temperature=0,
                max_tokens=100,
                top_p=1,
                frequency_penalty=0.0,
                presence_penalty=0.0,
                stop=["\n"],
            )
            out = response['choices'][0]['text']
            st.session_state['history'].append((input, out))
            # print(out)
            return out

        if st.session_state['ready']:
            if 'generated' not in st.session_state:
                st.session_state['generated'] = ["Welcome! You can now ask any questions"]

            if 'past' not in st.session_state:
                st.session_state['past'] = ["Hi Bot!"]

            # Container for chat history
            response_container = st.container()

            # Container for text box
            container = st.container()

            with container:
                with st.form(key='my_form', clear_on_submit=True):
                    user_input = st.text_input("Ask something:", placeholder="e.g: Summarize the paper in a few sentences", key='input')
                    submit_button = st.form_submit_button(label='Send')

                if submit_button and user_input:
                    output = await res(user_input)
                    st.session_state['past'].append(user_input)
                    st.session_state['generated'].append(output)

            if st.session_state['generated']:
                with response_container:
                    for i in range(len(st.session_state['generated'])):
                        message(st.session_state["past"][i], is_user=True, key=str(i) + '_user', avatar_style="thumbs")
                        message(st.session_state["generated"][i], key=str(i) + '🤖')
                        # , avatar_style="thumbs")👽🐧

asyncio.run(main())
